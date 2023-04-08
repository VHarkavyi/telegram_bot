import time
import pandas as pd
import telebot, datetime, csv, types
import csv
import openpyxl
import os
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove

max_attempts = 5
attempt_count = 0

while attempt_count < max_attempts:
    try:
        # token
        bot = telebot.TeleBot('6133056354:AAH7DGv0JOw1Tny3_DB7mcnbY4VHmBCPFNY')

        ############################################################################
        # Стартовыe кнопки
        keyboard_start = ReplyKeyboardMarkup(resize_keyboard=True)
        add_spend_button = KeyboardButton('Добавить трату')
        balance_button = KeyboardButton('Баланс')
        keyboard_start.add(add_spend_button)
        keyboard_start.add(balance_button)

        # Кнопки вернуться назад
        keyboard_back_to_start = ReplyKeyboardMarkup(resize_keyboard=True)
        back_to_start_button = KeyboardButton('Вернуться на старт')
        keyboard_back_to_start.add(back_to_start_button)

        # Кнопка пропуска + вернуться назад
        keyboard_skip = ReplyKeyboardMarkup(resize_keyboard=True)
        skip_button = KeyboardButton('Пропустить')
        keyboard_skip.add(skip_button)
        keyboard_skip.add(back_to_start_button)

        # кнопки категорий
        Groceries = KeyboardButton('Groceries')
        Lasure_Outing = KeyboardButton('Lasure/Outing')
        Javelina = KeyboardButton('Javelina')
        Homeware_Hygiene = KeyboardButton('Homeware + Hygiene')
        Utilities_Internet = KeyboardButton('Utilities + Internet')
        Subscriptions_phone = KeyboardButton('Subscriptions + Phone')
        Haircut = KeyboardButton('Haircut')
        Transport = KeyboardButton('Transport')
        Purchases = KeyboardButton('Purchases')
        Fox = KeyboardButton('Fox')
        Parents = KeyboardButton('Parents')
        Presents = KeyboardButton('Presents')
        Delivery = KeyboardButton('Delivery')
        Other = KeyboardButton('Other')
        Events = KeyboardButton('Events')
        Save = KeyboardButton('Save')

        keyboard_category = ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard_category.add(back_to_start_button)
        keyboard_category.add(Groceries)
        keyboard_category.add(Javelina)
        keyboard_category.add(Homeware_Hygiene)
        keyboard_category.add(Utilities_Internet)
        keyboard_category.add(Subscriptions_phone)
        keyboard_category.add(Haircut)
        keyboard_category.add(Transport)
        keyboard_category.add(Purchases)
        keyboard_category.add(Fox)
        keyboard_category.add(Parents)
        keyboard_category.add(Presents)
        keyboard_category.add(Delivery)
        keyboard_category.add(Other)
        keyboard_category.add(Events)
        keyboard_category.add(Save)

        # Кнопки источника
        MonoBlack = KeyboardButton('Mono Black')
        MonoWhite = KeyboardButton('Mono White')
        Cash = KeyboardButton('Cash')
        Fox = KeyboardButton('Fox')
        Ukrsib = KeyboardButton('Ukrsib')
        Privat = KeyboardButton('Privat')

        keyboard_source = ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard_source.add(back_to_start_button)
        keyboard_source.add(MonoBlack)
        keyboard_source.add(MonoWhite)
        keyboard_source.add(Cash)
        keyboard_source.add(Fox)
        keyboard_source.add(Ukrsib)
        keyboard_source.add(Privat)
        ####################################################################################

        user_global_state = {'step': 'start'}
        print(user_global_state)


        @bot.message_handler(commands=['start'])
        def start_handler(message):
            bot.send_message(message.chat.id, 'Выберите действие: ', reply_markup=keyboard_start)


        ################################################################################################
        # Check balance

        @bot.message_handler(func=lambda message: message.text == 'Баланс')
        def get_balance(message):
            if check_availability(message) == True:
                user_global_state[message.chat.id] = {'step': 'balance'}
                print(user_global_state)

                def send_balance(message):
                    # Получаем нужный лист
                    now = datetime.datetime.now()
                    data = now.strftime("%b %y")

                    df = pd.read_excel('Budget.xlsx', sheet_name=data, header=None)
                    # Получаем значение ячейки
                    value_mblack = df.at[23, 8]
                    value_mwhite = round(df.at[23, 9], 2)
                    value_cash = round(df.at[23, 10], 2)
                    value_ukrsib = round(df.at[23, 11], 2)
                    value_privat = round(df.at[23, 12], 2)

                    bot.send_message(message.chat.id,
                                     'MonoBlack = ' + str(value_mblack) + ' UAH\n'
                                     + 'MonoWhite = ' + str(value_mwhite) + ' UAH\n'
                                     + 'Cash = ' + str(value_cash) + ' UAH\n'
                                     + 'Ukrsib = ' + str(value_ukrsib) + ' UAH\n'
                                     + 'Privat = ' + str(value_privat) + ' UAH\n'
                                     , reply_markup=keyboard_start)

                    user_global_state[message.chat.id] = {'step': 'start'}
                    print(user_global_state)

                send_balance(message)


        #################################################################################################
        #   Add spends
        @bot.message_handler(func=lambda message: message.text == 'Добавить трату')
        def add_spend(message):
            if check_availability(message) == True:
                user_global_state[message.chat.id] = {'step': 'select_category'}
                print(user_global_state)
                bot.send_message(message.chat.id, 'Выберите категорию: ', reply_markup=keyboard_category)


        @bot.message_handler(
            func=lambda message: message.content_type == 'text' and user_global_state.get(message.chat.id, {}).get(
                'step') == 'select_category')
        def select_category(message):
            user_state = user_global_state.get(message.chat.id)
            if not user_state or user_state['step'] != 'select_category':
                return
            if message.text == 'Вернуться на старт':
                start_handler(message)
                user_global_state[message.chat.id] = {'step': 'start'}
                print(user_global_state)
            else:
                category = message.text
                user_global_state[message.chat.id]['category'] = category
                user_global_state[message.chat.id]['step'] = 'select_source'
                print(user_global_state)
                bot.send_message(message.chat.id, 'Выберите источник: ', reply_markup=keyboard_source)


        @bot.message_handler(
            func=lambda message: message.content_type == 'text' and user_global_state.get(message.chat.id, {}).get(
                'step') == 'select_source')
        def select_source(message):
            user_state = user_global_state.get(message.chat.id)
            if not user_state or user_state['step'] != 'select_source':
                return
            if message.text == 'Вернуться на старт':
                start_handler(message)
                user_global_state[message.chat.id] = {'step': 'start'}
                print(user_global_state)
            else:
                source = message.text
                user_global_state[message.chat.id]['source'] = source
                user_global_state[message.chat.id]['step'] = 'enter_amount'
                print(user_global_state)
                # bot.send_message(message.chat.id, 'Введите сумму:', reply_markup=ReplyKeyboardRemove())
                bot.send_message(message.chat.id, 'Введите сумму:', reply_markup=keyboard_back_to_start)


        @bot.message_handler(
            func=lambda message: message.content_type == 'text' and user_global_state.get(message.chat.id, {}).get(
                'step') == 'enter_amount')
        def enter_amount(message):
            user_state = user_global_state.get(message.chat.id)
            if not user_state or user_state['step'] != 'enter_amount':
                return
            if message.text == 'Вернуться на старт':
                start_handler(message)
                user_global_state[message.chat.id] = {'step': 'start'}
                print(user_global_state)
            else:
                try:
                    amount = float(message.text.replace(',', '.'))
                    user_global_state[message.chat.id]['amount'] = amount
                    user_global_state[message.chat.id]['step'] = 'enter_comment'
                    print(user_global_state)
                    bot.send_message(message.chat.id, f"Добавьте комментарий: ", reply_markup=keyboard_skip)
                except ValueError:
                    bot.send_message(message.chat.id, "Ошибка! Введите число.")


        @bot.message_handler(
            func=lambda message: message.content_type == 'text' and user_global_state.get(message.chat.id, {}).get(
                'step') == 'enter_comment')
        def add_comment(message):
            user_state = user_global_state.get(message.chat.id)
            if not user_state or user_state['step'] != 'enter_comment':
                return
            if message.text == 'Вернуться на старт':
                start_handler(message)
                user_global_state[message.chat.id] = {'step': 'start'}
                print(user_global_state)
            elif message.text == 'Пропустить':
                comment = ' '
                now = datetime.datetime.now()
                data = now.strftime("%d/%m")
                spend_entry = [user_state['category'], user_state['source'], int(user_state['amount']), data, comment]
                write_data_to_file(spend_entry, message)
                user_global_state[message.chat.id] = {'step': 'start'}
            else:
                comment = message.text
                now = datetime.datetime.now()
                data = now.strftime("%d/%m")
                spend_entry = [user_state['category'], user_state['source'], int(user_state['amount']), data, comment]
                write_data_to_file(spend_entry, message)
                user_global_state[message.chat.id] = {'step': 'start'}


        ###########################################################################################

        def write_data_to_file(spend_entry, message):
            with open('output.csv', mode='w', encoding='utf-8', newline='') as file:
                writer = csv.writer(file, delimiter=';', quoting=csv.QUOTE_MINIMAL)
                writer.writerow(spend_entry)
                file.close()

                def write_data_to_excel():
                    # Открываем файл CSV и читаем его содержимое
                    with open('output.csv', 'r', encoding='utf-8') as csv_file:
                        csv_reader = csv.reader(csv_file)
                        csv_data = list(csv_reader)
                        print(csv_data)

                    # Открываем книгу Excel
                    workbook = openpyxl.load_workbook('Budget.xlsx')

                    # Получаем нужный лист
                    now = datetime.datetime.now()
                    data = now.strftime("%b %y")
                    worksheet = workbook[data]

                    # Находим последнюю заполненную строку в таблице
                    last_row = 5
                    while worksheet.cell(row=last_row, column=17).value is not None:
                        last_row += 1

                    # Записываем данные в Excel-файл, начиная со следующей свободной ячейки в столбце A
                    for row in csv_data:
                        row_data = row[0].split(';')
                        for col_index, cell_value in enumerate(row_data):
                            if cell_value.isnumeric():
                                worksheet.cell(row=last_row, column=col_index + 17, value=int(cell_value))
                            else:
                                worksheet.cell(row=last_row, column=col_index + 17, value=cell_value)
                        last_row += 1

                    # Сохраняем изменения в книге Excel
                    workbook.save('Budget.xlsx')
                    workbook.close()

                write_data_to_excel()
                bot.send_message(message.chat.id, "Спасибо, ваша трата успешно записана! Выберите следующее действие.",
                                 reply_markup=keyboard_start)
                os.remove('output.csv')
                user_global_state[message.chat.id] = {'step': 'start'}
                print(user_global_state)


        def check_availability(message):
            try:
                workbook = openpyxl.load_workbook('Budget.xlsx')
                workbook.save('Budget.xlsx')
                workbook.close()
                return True
            except PermissionError:
                bot.send_message(message.chat.id, "Файл открыт другой альпаськой, попробуй позже!")
                user_global_state[message.chat.id] = {'step': 'start'}


        ###########################################################################################

        # Запуск бота
        bot.polling(none_stop=True)
        attempt_count = 0
    except Exception as e:
        print(f"Ошибка: {e}. Перезапуск через 5 секунд.")
        time.sleep(5)
        attempt_count += 1

print("Достигнуто максимальное количество попыток. Скрипт остановлен.")
